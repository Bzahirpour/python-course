import openpyxl as xl
from openpyxl.chart import BarChart, Reference

print(xl.__version__)

wb = xl.load_workbook('transactions.xlsx')
sheet1 = wb['Sheet1']
cella1 = sheet1['a1']
print(cella1.value)

print(sheet1.max_row)

for row in range(2, sheet1.max_row + 1):
    cell = sheet1.cell(row, 3)
    print(cell.value)
    corrected_price = cell.value * .1
    print(corrected_price)
    corrected_price_cell = sheet1.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet1, min_row=2, max_row=sheet1.max_row, min_col=4, max_col=4)
chart = BarChart()
chart.add_data(values)
sheet1.add_chart(chart, 'e2')
wb.save('transations2.xlsx')


# converted to a function
def update_prices(filename):
    wb = xl.load_workbook('filename')
    sheet1 = wb['Sheet1']

    for row in range(2, sheet1.max_row + 1):
        cell = sheet1.cell(row, 3)
        corrected_price = cell.value * .1
        corrected_price_cell = sheet1.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet1, min_row=2, max_row=sheet1.max_row, min_col=4, max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet1.add_chart(chart, 'e2')
    wb.save('filename.xlsx')